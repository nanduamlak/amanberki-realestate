import sys
import os

# Add cPanel local user site-packages to path so openpyxl is always found on production
sys.path.insert(0, "/home/amanbeqj/.local/lib/python3.9/site-packages")

import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    # Read plots from stdin
    try:
        plots = json.load(sys.stdin)
    except Exception as e:
        print(f"Error reading JSON from stdin: {e}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    ws = wb.create_sheet(title="Sheet2")
    
    # Enable grid lines visible
    ws.views.sheetView[0].showGridLines = True

    # 1. Header row
    headers = [
        "Block No.", "Plot No. ", "Plot Size (Sq.M)", "Built-up Area", 
        " Purchaser Name ", "Title Deeds Status", "Contractor", "Construction Status", "Remark\\Comments"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # Style: bold, font size 10, center aligned, lime green background (#92D050)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.row_dimensions[1].height = 25
    
    # Styles for data rows
    thin_border = Border(
        left=Side(style='thin', color='A6A6A6'),
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='A6A6A6'),
        bottom=Side(style='thin', color='A6A6A6')
    )
    
    # Group plots by block
    block_groups = {}
    for p in plots:
        b = p.get("blockNumber", 0)
        if b not in block_groups:
            block_groups[b] = []
        block_groups[b].append(p)
          
    # Sort blocks
    sorted_blocks = sorted(block_groups.keys())
    
    current_row = 2
    
    for block_idx, block_num in enumerate(sorted_blocks):
        # Add divider row between blocks
        if block_idx > 0:
            # Replicate the grey divider row exactly: bg-color #7F7F7F, height 12, merged A:I
            ws.row_dimensions[current_row].height = 12
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            for col_idx in range(1, 10):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
                cell.border = thin_border
            current_row += 1
            
        block_plots = block_groups[block_num]
        
        # Sort plots: numeric primarily
        def get_plot_key(x):
            pn = x.get("plotNumber", "")
            try:
                return (0, int(pn))
            except ValueError:
                return (1, pn)
        block_plots = sorted(block_plots, key=get_plot_key)
        
        start_row = current_row
        
        for plot_idx, p in enumerate(block_plots):
            # Only write block label on first row of block group (merging will cover it)
            val_block = ""
            if plot_idx == 0:
                val_block = p.get("blockLabel") or str(p.get("blockNumber", ""))
                
            row_data = [
                val_block,
                p.get("plotNumber", ""),
                p.get("plotSize", ""),
                p.get("builtArea", ""),
                p.get("purchaserName", ""),
                p.get("titleDeedsStatus", ""),
                p.get("contractorName", ""),
                p.get("constructionStatus", ""),
                p.get("remark", "")
            ]
            
            # Dynamically calculate row height based on Remark length (column width is 45)
            remark_text = p.get("remark") or ""
            lines = max(1, (len(str(remark_text)) + 40) // 45)
            ws.row_dimensions[current_row].height = max(20, lines * 15)
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                
                # Alignment
                if col_idx == 9:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx in [1, 2, 3, 6]: # Block, Plot, Size, Deed Status
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                # Format plot size as number if possible
                if col_idx == 3 and val != "":
                    try:
                        cell.value = float(val)
                        cell.number_format = "#,##0"
                    except ValueError:
                        pass
                        
            current_row += 1
            
        end_row = current_row - 1
        
        # Merge Block No. cells
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            # Center the merged block label
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            
    # Auto-adjust column widths
    column_widths = {
        1: 12, # Block No.
        2: 10, # Plot No.
        3: 16, # Plot Size
        4: 14, # Built-up Area
        5: 28, # Purchaser Name
        6: 20, # Title Deeds Status
        7: 18, # Contractor
        8: 24, # Construction Status
        9: 45  # Remark
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    # Save output to path provided as argument
    if len(sys.argv) > 1:
        wb.save(sys.argv[1])
    else:
        wb.save("temp_export.xlsx")

if __name__ == "__main__":
    main()

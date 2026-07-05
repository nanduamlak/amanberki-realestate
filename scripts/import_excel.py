"""
Import 'Tulu Dimtu Inventory last final.xlsx' -> local PostgreSQL DB
Upserts plot_details rows then refreshes properties aggregates.
Run: python scripts/import_excel.py
"""

import openpyxl
import psycopg2

# ── DB connection ───────────────────────────────────────────────────────────────
DB = dict(host="localhost", port=5432, dbname="real_estate_db",
          user="postgres", password="123456")

EXCEL      = "Tulu Dimtu Inventory last final.xlsx"
SHEET      = "Sheet2"
DATA_START = 12   # first data row (row 11 = headers)

COMPANY_NAMES = {"tulu dimtu real estate", "tulu dimtu", ""}

def normalise_deed(v):
    if not v: return ""
    v = str(v).strip().upper()
    if v == "ISSUED":     return "ISSUED"
    if v == "PENDING":    return "PENDING"
    if v == "NOT ISSUED": return "NOT ISSUED"
    if v == "TITLE DEEDS STATUS": return ""
    return str(v)

def load_excel():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    blocks = {}
    current_block = None
    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row, values_only=True):
        # Column A = block number (only on the first row of each block)
        if row[0] is not None and isinstance(row[0], (int, float)):
            current_block = int(row[0])
            blocks.setdefault(current_block, [])
        if current_block is None:
            continue
        plot_no = row[1]
        if plot_no is None:
            continue
        try:
            plot_no = int(plot_no)
        except (ValueError, TypeError):
            continue

        plot_size  = str(row[2]).strip() if row[2] is not None else "0"
        purchaser  = str(row[4]).strip() if row[4] else ""
        deed       = normalise_deed(row[5])
        contractor = str(row[6]).strip() if row[6] else None
        const_st   = str(row[7]).strip() if row[7] else ""
        remark     = str(row[8]).strip() if row[8] else ""

        blocks[current_block].append({
            "plot_no":      str(plot_no),
            "plot_size":    plot_size,
            "purchaser":    purchaser,
            "deed":         deed,
            "contractor":   contractor,
            "const_status": const_st,
            "remark":       remark,
        })
    return blocks

def get_block_id(cur, block_no: int):
    cur.execute("SELECT id FROM properties WHERE block_number = %s", (block_no,))
    row = cur.fetchone()
    return row[0] if row else None

def refresh_block(cur, block_id: str):
    """Re-aggregate plot_details -> properties for one block."""
    # Count sold vs active based on purchaser name
    cur.execute("""
        UPDATE properties SET
            no_of_plots  = (SELECT COUNT(*) FROM plot_details WHERE block_id = %s),
            sold_plots   = (SELECT COUNT(*) FROM plot_details
                            WHERE block_id = %s
                              AND purchaser_name IS NOT NULL
                              AND LOWER(TRIM(purchaser_name)) NOT IN ('tulu dimtu real estate','tulu dimtu','')),
            active_plots = (SELECT COUNT(*) FROM plot_details
                            WHERE block_id = %s
                              AND (purchaser_name IS NULL
                                   OR LOWER(TRIM(purchaser_name)) IN ('tulu dimtu real estate','tulu dimtu',''))),
            area         = COALESCE((
                             SELECT SUM(
                               CASE WHEN plot_size ~ '^[0-9]+\\+[0-9]+$'
                                    THEN (SPLIT_PART(plot_size,'+',1)::numeric +
                                          SPLIT_PART(plot_size,'+',2)::numeric)
                                    WHEN plot_size ~ '^[0-9]+(\\.[0-9]+)?$'
                                    THEN plot_size::numeric
                                    ELSE 0 END
                             ) FROM plot_details WHERE block_id = %s
                           ), 0)
        WHERE id = %s
    """, (block_id, block_id, block_id, block_id, block_id))

def main():
    print("Loading Excel ...")
    blocks = load_excel()
    print(f"  Found {len(blocks)} blocks, "
          f"{sum(len(v) for v in blocks.values())} plots total")

    conn = psycopg2.connect(**DB)
    conn.autocommit = False
    cur  = conn.cursor()

    inserted = updated = skipped = 0

    for block_no, plots in sorted(blocks.items()):
        block_id = get_block_id(cur, block_no)
        if block_id is None:
            print(f"  SKIP Block {block_no:>2} — not found in properties table")
            skipped += 1
            continue

        for p in plots:
            plot_number = p["plot_no"]
            cur.execute(
                "SELECT id FROM plot_details WHERE block_id = %s AND plot_number = %s",
                (block_id, plot_number)
            )
            existing = cur.fetchone()

            if existing:
                cur.execute("""
                    UPDATE plot_details SET
                        plot_size           = %s,
                        purchaser_name      = %s,
                        title_deeds_status  = %s,
                        contractor_name     = %s,
                        construction_status = %s,
                        remark              = %s,
                        updated_at          = NOW()
                    WHERE block_id = %s AND plot_number = %s
                """, (
                    p["plot_size"],
                    p["purchaser"],
                    p["deed"],
                    p["contractor"],
                    p["const_status"],
                    p["remark"],
                    block_id, plot_number
                ))
                updated += 1
            else:
                cur.execute("""
                    INSERT INTO plot_details
                        (block_id, plot_number, plot_size,
                         purchaser_name, title_deeds_status,
                         contractor_name, construction_status, remark)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    block_id, plot_number,
                    p["plot_size"],
                    p["purchaser"],
                    p["deed"],
                    p["contractor"],
                    p["const_status"],
                    p["remark"],
                ))
                inserted += 1

        refresh_block(cur, block_id)
        print(f"  Block {block_no:>2}: {len(plots)} plots  "
              f"(ins={sum(1 for p in plots if p)})")

    conn.commit()
    cur.close()
    conn.close()

    print(f"\nDone.")
    print(f"  Inserted : {inserted}")
    print(f"  Updated  : {updated}")
    print(f"  Skipped  : {skipped} blocks not in DB")

if __name__ == "__main__":
    main()

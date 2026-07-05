"""
Convert PPK v3 → OpenSSH PEM using puttygen, then sync Excel data to cPanel via SSH tunnel.
"""
import subprocess, sys, os, io, tempfile
import openpyxl
import psycopg2

PPK_PATH   = r"C:\Users\andua\Downloads\id_rsa (7).ppk"
PASSPHRASE = "nandu25588032,"
PEM_PATH   = r"C:\Users\andua\Downloads\id_rsa_temp.pem"
SSH_HOST   = "91.204.209.50"
SSH_PORT   = 22
SSH_USER   = "amanbeqj"

REMOTE_DB = dict(
    dbname   = "amanbeqj_realestatedb",
    user     = "amanbeqj_dbuser",
    password = "Nandu25588032,",
)

EXCEL      = "Tulu Dimtu Inventory last final.xlsx"
SHEET      = "Sheet2"
DATA_START = 12

def convert_ppk_to_pem():
    """Use puttygen.exe to convert PPK → OpenSSH PEM."""
    puttygen = r"C:\Program Files\PuTTY\puttygen.exe"
    result = subprocess.run(
        [puttygen, PPK_PATH,
         f"--old-passphrase", PASSPHRASE,
         "-O", "private-openssh",
         "-o", PEM_PATH],
        capture_output=True, text=True, timeout=15
    )
    if not os.path.exists(PEM_PATH):
        # Try without --old-passphrase flag (older puttygen syntax uses -P)
        result = subprocess.run(
            [puttygen, PPK_PATH,
             "-P", PASSPHRASE,
             "-O", "private-openssh",
             "-o", PEM_PATH],
            capture_output=True, text=True, timeout=15
        )
    if os.path.exists(PEM_PATH):
        print(f"  PEM created: {PEM_PATH}")
        return True
    print(f"  puttygen stdout: {result.stdout}")
    print(f"  puttygen stderr: {result.stderr}")
    return False

def normalise_deed(v):
    if not v: return ""
    v = str(v).strip().upper()
    if v == "ISSUED":             return "ISSUED"
    if v == "PENDING":            return "PENDING"
    if v == "NOT ISSUED":         return "NOT ISSUED"
    if v == "TITLE DEEDS STATUS": return ""
    return str(v)

def load_excel():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    blocks = {}
    current_block = None
    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row, values_only=True):
        if row[0] is not None and isinstance(row[0], (int, float)):
            current_block = int(row[0])
            blocks.setdefault(current_block, [])
        if current_block is None:
            continue
        plot_no = row[1]
        if plot_no is None:
            continue
        try:
            plot_no = int(plot_no)
        except (ValueError, TypeError):
            continue
        blocks[current_block].append({
            "plot_no":      str(plot_no),
            "plot_size":    str(row[2]).strip() if row[2] is not None else "0",
            "purchaser":    str(row[4]).strip() if row[4] else "",
            "deed":         normalise_deed(row[5]),
            "contractor":   str(row[6]).strip() if row[6] else None,
            "const_status": str(row[7]).strip() if row[7] else "",
            "remark":       str(row[8]).strip() if row[8] else "",
        })
    return blocks

def refresh_block(cur, block_id):
    cur.execute("""
        UPDATE properties SET
            no_of_plots  = (SELECT COUNT(*) FROM plot_details WHERE block_id = %s),
            sold_plots   = (SELECT COUNT(*) FROM plot_details
                            WHERE block_id = %s
                              AND purchaser_name IS NOT NULL
                              AND LOWER(TRIM(purchaser_name)) NOT IN ('tulu dimtu real estate','tulu dimtu','')),
            active_plots = (SELECT COUNT(*) FROM plot_details
                            WHERE block_id = %s
                              AND (purchaser_name IS NULL
                                   OR LOWER(TRIM(purchaser_name)) IN ('tulu dimtu real estate','tulu dimtu',''))),
            area         = COALESCE((
                             SELECT SUM(
                               CASE WHEN plot_size ~ '^[0-9]+\\+[0-9]+$'
                                    THEN (SPLIT_PART(plot_size,'+',1)::numeric +
                                          SPLIT_PART(plot_size,'+',2)::numeric)
                                    WHEN plot_size ~ '^[0-9]+(\\.[0-9]+)?$'
                                    THEN plot_size::numeric
                                    ELSE 0 END
                             ) FROM plot_details WHERE block_id = %s
                           ), 0)
        WHERE id = %s
    """, (block_id, block_id, block_id, block_id, block_id))

def main():
    print("Loading Excel ...")
    blocks = load_excel()
    print(f"  {len(blocks)} blocks, {sum(len(v) for v in blocks.values())} plots")

    # Step 1: Convert PPK → PEM
    print(f"\nConverting PPK to OpenSSH PEM ...")
    if not convert_ppk_to_pem():
        print("ERROR: Could not convert PPK to PEM. Check puttygen is installed.")
        return

    # Step 2: Open SSH tunnel in background process
    print(f"\nOpening SSH tunnel to {SSH_HOST}:5432 ...")
    ssh_cmd = [
        r"C:\Windows\System32\OpenSSH\ssh.exe",
        "-i", PEM_PATH,
        "-o", "StrictHostKeyChecking=no",
        "-o", "BatchMode=yes",
        "-N",
        "-L", "5433:127.0.0.1:5432",
        f"{SSH_USER}@{SSH_HOST}",
        "-p", str(SSH_PORT)
    ]
    tunnel_proc = subprocess.Popen(ssh_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

    import time
    time.sleep(6)  # Give tunnel time to establish

    if tunnel_proc.poll() is not None:
        # Tunnel already exited
        out, err = tunnel_proc.communicate()
        print(f"ERROR: Tunnel process exited early")
        print(f"  stdout: {out.decode()}")
        print(f"  stderr: {err.decode()}")
        return

    print(f"  Tunnel process running (PID {tunnel_proc.pid})")

    try:
        # Step 3: Connect and sync
        print("Connecting to cPanel DB via tunnel ...")
        conn = psycopg2.connect(
            host="127.0.0.1",
            port=5433,
            **REMOTE_DB
        )
        conn.autocommit = False
        cur = conn.cursor()
        print("  Connected!")

        inserted = updated = skipped = 0

        for block_no, plots in sorted(blocks.items()):
            cur.execute("SELECT id FROM properties WHERE block_number = %s", (block_no,))
            row = cur.fetchone()
            if not row:
                print(f"  SKIP Block {block_no:>2}")
                skipped += 1
                continue
            block_id = row[0]

            for p in plots:
                cur.execute(
                    "SELECT id FROM plot_details WHERE block_id=%s AND plot_number=%s",
                    (block_id, p["plot_no"])
                )
                if cur.fetchone():
                    cur.execute("""
                        UPDATE plot_details SET
                            plot_size=(%s), purchaser_name=(%s), title_deeds_status=(%s),
                            contractor_name=(%s), construction_status=(%s), remark=(%s),
                            updated_at=NOW()
                        WHERE block_id=(%s) AND plot_number=(%s)
                    """, (p["plot_size"], p["purchaser"], p["deed"],
                          p["contractor"], p["const_status"], p["remark"],
                          block_id, p["plot_no"]))
                    updated += 1
                else:
                    cur.execute("""
                        INSERT INTO plot_details
                            (block_id, plot_number, plot_size, purchaser_name,
                             title_deeds_status, contractor_name, construction_status, remark)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (block_id, p["plot_no"], p["plot_size"], p["purchaser"],
                          p["deed"], p["contractor"], p["const_status"], p["remark"]))
                    inserted += 1

            refresh_block(cur, block_id)
            print(f"  Block {block_no:>2}: {len(plots)} plots")

        conn.commit()
        cur.close()
        conn.close()
        print(f"\n✅ cPanel DB updated!")
        print(f"  Inserted : {inserted}")
        print(f"  Updated  : {updated}")
        print(f"  Skipped  : {skipped}")

    finally:
        tunnel_proc.terminate()
        if os.path.exists(PEM_PATH):
            os.remove(PEM_PATH)
        print("Tunnel closed and temp key removed.")

if __name__ == "__main__":
    main()
